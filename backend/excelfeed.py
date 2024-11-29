import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import time
from datetime import datetime
import os

ticker = 'INFY'
url = f'https://www.google.com/finance/quote/{ticker}:NSE'
file_name = "stock_prices.xlsx"

def is_file_open(filepath):
    try:
        with open(filepath, 'a'):
            pass
        return False
    except IOError:
        return True

try:
    # Load or create Excel file
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = "Timestamp"
        sheet['B1'] = "Stock Symbol"
        sheet['C1'] = "Stock Price"

    for i in range(3):  # Loop to fetch data
        response = requests.get(url)
        if response.status_code != 200:
            print(f"Failed to fetch webpage: {response.status_code}")
            break

        soup = BeautifulSoup(response.text, 'html.parser')

        try:
            stock_price = soup.find('div', class_='YMlKec fxKbKc').text
            print(f"Stock Price: {stock_price}")
        except AttributeError:
            print("Failed to locate stock price.")
            break

        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sheet.append([current_time, f"{ticker}:NSE", stock_price])

        if is_file_open(file_name):
            print(f"Error: '{file_name}' is open. Please close it.")
            break
        else:
            workbook.save(file_name)
            print(f"Stock price saved to '{file_name}'.")

        time.sleep(10)

except Exception as e:
    print(f"An error occurred: {e}")
